import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the meals database
with open('src/data/mealsData.json', 'r') as f:
    data = json.load(f)

meals = data['meals']

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create sheets for different meal types
meal_types = {
    'All Meals': meals,
    'Breakfast': [m for m in meals if 'breakfast' in m['tags']],
    'Lunch': [m for m in meals if 'lunch' in m['tags']],
    'Dinner': [m for m in meals if 'dinner' in m['tags']],
    'Snacks': [m for m in meals if 'snack' in m['tags']],
}

# Weather-based sheets
weather_types = {
    'Hot Weather': [m for m in meals if 'hot' in m.get('weatherCategories', [])],
    'Cold Weather': [m for m in meals if 'cold' in m.get('weatherCategories', [])],
    'Rainy Weather': [m for m in meals if 'rainy' in m.get('weatherCategories', [])],
    'Humid Weather': [m for m in meals if 'humid' in m.get('weatherCategories', [])],
    'Cloudy Weather': [m for m in meals if 'cloudy' in m.get('weatherCategories', [])],
}

# Diet-based sheets
diet_types = {
    'Vegetarian': [m for m in meals if m.get('type') == 'veg'],
    'Non-Vegetarian': [m for m in meals if m.get('type') == 'non-veg'],
}

# Combine all sheet definitions
all_sheets = {**meal_types, **weather_types, **diet_types}

# Column headers
headers = ['ID', 'Meal Name', 'Description', 'Type', 'Category', 'Weather', 'Ingredients', 
           'Protein', 'Calories', 'Prep Time', 'Emoji', 'Image URL']

def create_sheet(wb, sheet_name, meals_list):
    """Create a formatted sheet with meal data"""
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_num, meal in enumerate(meals_list, 2):
        ws.cell(row=row_num, column=1, value=meal['id']).border = border
        ws.cell(row=row_num, column=2, value=meal['name']).border = border
        ws.cell(row=row_num, column=3, value=meal['description']).border = border
        ws.cell(row=row_num, column=4, value=meal['type'].upper()).border = border
        ws.cell(row=row_num, column=5, value=', '.join(meal['tags']).title()).border = border
        ws.cell(row=row_num, column=6, value=', '.join(meal.get('weatherCategories', [])).title()).border = border
        ws.cell(row=row_num, column=7, value=', '.join(meal['ingredients'])).border = border
        ws.cell(row=row_num, column=8, value=meal['protein']).border = border
        ws.cell(row=row_num, column=9, value=meal['calories']).border = border
        ws.cell(row=row_num, column=10, value=meal['prepTime']).border = border
        ws.cell(row=row_num, column=11, value=meal['image']).border = border
        
        # Add hyperlink to image URL
        image_cell = ws.cell(row=row_num, column=12, value=meal.get('imageUrl', ''))
        image_cell.hyperlink = meal.get('imageUrl', '')
        image_cell.font = Font(color="0563C1", underline="single")
        image_cell.border = border
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # ID
        'B': 30,  # Meal Name
        'C': 35,  # Description
        'D': 12,  # Type
        'E': 15,  # Category
        'F': 25,  # Weather
        'G': 40,  # Ingredients
        'H': 10,  # Protein
        'I': 10,  # Calories
        'J': 12,  # Prep Time
        'K': 8,   # Emoji
        'L': 70,  # Image URL
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_num in range(2, len(meals_list) + 2):
        if row_num % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = light_fill

# Create all sheets
for sheet_name, meals_list in all_sheets.items():
    create_sheet(wb, sheet_name, meals_list)

# Create a summary sheet at the beginning
summary_ws = wb.create_sheet(title="📊 Summary", index=0)

# Summary data
summary_data = [
    ["FitFusion Meal Database", ""],
    ["Total Meals", len(meals)],
    ["", ""],
    ["By Meal Type", ""],
    ["Breakfast", len([m for m in meals if 'breakfast' in m['tags']])],
    ["Lunch", len([m for m in meals if 'lunch' in m['tags']])],
    ["Dinner", len([m for m in meals if 'dinner' in m['tags']])],
    ["Snacks", len([m for m in meals if 'snack' in m['tags']])],
    ["", ""],
    ["By Diet Type", ""],
    ["Vegetarian", len([m for m in meals if m.get('type') == 'veg'])],
    ["Non-Vegetarian", len([m for m in meals if m.get('type') == 'non-veg'])],
    ["", ""],
    ["By Weather", ""],
    ["Hot Weather", len([m for m in meals if 'hot' in m.get('weatherCategories', [])])],
    ["Cold Weather", len([m for m in meals if 'cold' in m.get('weatherCategories', [])])],
    ["Rainy Weather", len([m for m in meals if 'rainy' in m.get('weatherCategories', [])])],
    ["Humid Weather", len([m for m in meals if 'humid' in m.get('weatherCategories', [])])],
    ["Cloudy Weather", len([m for m in meals if 'cloudy' in m.get('weatherCategories', [])])],
    ["", ""],
    ["Features", ""],
    ["✅ All meals have high-quality images", ""],
    ["✅ Images from Unsplash food photography", ""],
    ["✅ Clickable image URLs in each sheet", ""],
    ["✅ Organized by meal type, weather, and diet", ""],
]

# Write summary
for row_num, (label, value) in enumerate(summary_data, 1):
    cell_a = summary_ws.cell(row=row_num, column=1, value=label)
    cell_b = summary_ws.cell(row=row_num, column=2, value=value)
    
    if row_num == 1:
        cell_a.font = Font(bold=True, size=16, color="4472C4")
    elif label in ["By Meal Type", "By Diet Type", "By Weather", "Features"]:
        cell_a.font = Font(bold=True, size=12, color="4472C4")
    elif label == "Total Meals":
        cell_a.font = Font(bold=True)
        cell_b.font = Font(bold=True, size=14, color="FF0000")

summary_ws.column_dimensions['A'].width = 40
summary_ws.column_dimensions['B'].width = 15

# Save the workbook
filename = 'FitFusion_Meals_Database_220.xlsx'
wb.save(filename)

print(f"✅ Excel file created successfully: {filename}")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
print(f"📝 Total meals: {len(meals)}")
print(f"\nSheets created:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
